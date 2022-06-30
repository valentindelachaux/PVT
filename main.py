import math
from datetime import datetime
import openpyxl as opxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import model as ty

import heat_transfer as bht

import os

par = {}

## Physics constants

par["sigma"] = 1 # Stefan-Boltzmann constant

## PV

par["eta_nom"] = 1 # nominal efficiency of PV panel

par["Eff_T"] = 1
par["T_ref"] = 25 # reference temperature, often 25°C

par["Eff_G"] = 1
par["G_ref"] = 1000 # reference irradiance, often 1000 W/m2

#par["X_rad"] = 1
par["X_corr"] = 1 

## Heat exchanger specs

par["tube_conv"] = 1

par["A_G"] = 1

par["W"] = 1 # the width (x-direction) between adjacent fluid tubes
par["D_tube"] = 1 #
par["l_c"] = 1
par["l_B"] = 1
par["L_af"] = 1  
par["iota"] = 1
par["N_meander"] = 1
par["N_harp"] = 1 # number of identical tubes carrying fluid through the collector
par["N_harp_actual"] = 1 # actual numer of identical tubes in parallel (harp geometry)
par["L_riser"] = 1 # the length of the collector along the flow direction = L_riser in Excel

par["D"] = 1
par["D_4"] = 1
par["l_i"] = 1

par["insulated"] = 1

## Additionnal fins = ailettes

par["ailette"] = 1

par["geometry"] = 1

par["fin_0"] = 1
par["N_f0"] = 1
par["L_f0"] = 1
par["delta_f0"] = 1

par["fin_1"] = 1
par["N_f1"] = 1
par["L_f1"] = 1
par["delta_f1"] = 1
par["delta_f1_int"] = 1
par["coeff_f1"] = 1

par["fin_2"] = 1
par["N_f2"] = 1
par["L_f2"] = 1
par["delta_f2"] = 1

par["fin_3"] = 1
par["N_f3"] = 1
par["L_f3"] = 1
par["delta_f3"] = 1

par["delta"] = 1
par["L_a"] = 1

par["Heta"] = 1

par["N_ail"] = 1
par["DELTA_a"] = 1

## Thermal / radiance

par["tau_alpha"] = 1 # transmittance-absorptance product for the solar collector
par["eps"] = 1 # emissivity of the top surface of the collector (PV surface)

# Geometry and thermal conductivities

par["k_air"] = 1
par["air_layer"] = 1

par["k_abs"] = 1 # thermal conductivity of the plate material
par["lambd_abs"] = 1 # thickness of the absorber plate

par["lambd_riser"] = 1
par["k_riser"] = 1

par["k_ail"] = 1 # thermal conductivity of the fin
par["lambd_ail"] = 1 # thickness of the fin

par["k_insulation"] = 1
par["e_insulation"] = 1

par["h_fluid"] = 1 # internal fluid heat transfer coefficient
par["h_top"] = 1
par["a_htop"] = 1
par["b_htop"] = 1 
par["coeff_h_top"] = 1
par["coeff_h_back"] = 1

par["h_inner"] = 1

par["R_TOP"] = 1 # instead of 1/h_outer in the document
par["R_INTER"] = 1 # = R_1 = R_INTER = resistance to heat transfer from the PV cells to the absorber plate
par["R_2"] = 1
par["R_B"] = 1 # resistance from the absorber through the back of the collector
par["C_B"] = 1 # the conductance between the absorber plate and the bonded tube

## Meteo

par["G_T0"] = 1 # total solar radiation (beam + diffuse) incident upon the collector surface = POA irradiance
par["G_p"] = 1 # infra-red 
par["coeff_G_p"] = 1
par["T_sky"] = 1 # sky temperature for long-wave radiation calculations
par["T_amb"] = 1 
par["T_back"] = 1
par["u"] = 1 # wind speed

## Fluid

par["T_fluid_in0"] = 1
par["C_p"] = 1 # specific heat of the fluid flowing through the PV/T collector
par["m_dot"] = 1 # flow rate of fluid through the solar collector

par["k_fluid"] = 1
par["rho_fluid"] = 1
par["mu_fluid"] = 1


## Installation

par["theta"] = 1 # angle of incidence

## Type de test

par["test"] = 1

## Keys  

list_parameters = []
*list_parameters, = par

# "sigma", "eta_nom","Eff_T","T_ref","Eff_G","G_ref","X_corr","W","D_tube","N_harp",
# "L_riser","tau_alpha","eps","_abs","lambd_abs","h_fluid","h_top","R_TOP","R_INTER","R_B","C_B","G_T0",
# "G_p","T_sky","T_amb","T_back","u", "T_fluid_in0","C_p","m_dot","theta"

## Excel parameters

path = os.getcwd()
path = r'C:\Users\valen\Documents\DualSun\1Dmodel'
print(path)

inp = r'\Inputs.xlsx'
fichier_i = path+inp
wbi = opxl.load_workbook(fichier_i,data_only=True)
sheet_i = wbi["Main"]

## Initialize paramters from Excel file
for i in range(len(list_parameters)):
    nom_var = list_parameters[i]
    cell = ty.find_cell_by_name(wbi,nom_var)
    valeur = sheet_i[cell].value
    
    par[nom_var]=valeur

wbi.close()

# Calculate X_rad which depends on G_T0
ty.X_rad(par)

#Calculate the conductance between the absorber and the fluid through any riser
ty.C_B(par)

# Excel file for outputs

now = datetime.now()
heure = str(now.hour)
minute = str(now.minute)
seconde = str(now.second)
moment = heure+'-'+minute+'-'+seconde

wbo = opxl.Workbook()
sheet_o = wbo.active
sheet_o.title = "Inputs"

## Test profile ##########################################################################################################################################""

##### Simulations with different meteos

G_list = [966]
#G_list = [0]
#G_list = [1000,1000]
coeff_G_p_list = [0]
#G_p_list = [0]
u_list = [0,0.7,1.4,2.7]
#u_list = [0]
T_amb_list = [265,270,275,280,285,290,295,300,305]
#T_amb_list = [280,300]

#compt = 2

T_f_in_list = [263,268,273,278,283,288,293,298,303,308,313,318,323]

T_guess_list = [293]


# Pour TUV

cond = r'\220321_TUV_test_conditions.xlsx'
fichier = path+cond
condi = pd.read_excel(fichier,"Non insulated v4 TUV")
condi.drop(index=condi.index[0], 
    axis=0, 
    inplace=True)

## Dossier et fichier

pathout = path+'\\Outputs'

exist = os.path.exists(pathout)

if exist == False:
    os.makedirs(pathout)

flag = 1

# copié dans l'ipynb
if flag == 1:

    if par["test"] == "TUV":
        df_par,df,X = ty.test_condi_list(par,condi)
    else: # general
        df_par,df,X = ty.test_meander_condi(par,G_list,coeff_G_p_list,T_amb_list,u_list,T_guess_list,T_f_in_list)

    # Creating Excel outputs
    if par["test"] == "TUV":
        prefixe = r'\OutputsTUV'
    else: # general
        prefixe = r'\Parametric'

    suffixe = prefixe+str(par["ailette"])+r'-'+moment+r'.xlsx'
    fichier_o = pathout+suffixe

    for d in dataframe_to_rows(df_par, index=True, header=False):
        sheet_o.append(d)

    wso = wbo.create_sheet("Outputs")

    for d in dataframe_to_rows(df, index=True, header=True):
        wso.append(d)

    huit = len(X[0])-1

    index_coeff = ['a1','a2','a3','a4','a6','a7','a8']
    ws_lsm = wbo.create_sheet("LSM")
    ws_lsm['A'+str(1)] = 'a0'
    ws_lsm['B'+str(1)] = X[0][len(X[0])-1]

    print('a0 : ',X[0][len(X[0])-1])

    ws_lsm['C'+str(1)] = X[3][len(X[0])-1]
    for l in range(huit):
        ws_lsm['A'+str(l+2)] = index_coeff[l]
        if l != 3:
            ws_lsm['B'+str(l+2)] = -X[0][l]
            print(index_coeff[l],' : ',-X[0][l])

        else:
            ws_lsm['B'+str(l+2)] = X[0][l]
            print(index_coeff[l],' : ',X[0][l])

        ws_lsm['C'+str(l+2)] = X[3][l]

    ws_lsm['A'+str(huit+2)] = 'a0 (DS) at 1,3 m/s'
    ws_lsm['B'+str(huit+2)] = X[0][len(X[0])-1] - (-X[0][4]*(1.3-3))
    print(round(X[0][len(X[0])-1] - (-X[0][4]*(1.3-3)),3)*100,'%')

    ws_lsm['A'+str(huit+3)] = 'a1 (DS) at 1,3 m/s'
    ws_lsm['B'+str(huit+3)] = -X[0][0] - X[0][2]*(1.3-3)
    print(round(-X[0][0] - X[0][2]*(1.3-3),1))

    wbo.save(filename = fichier_o)
    wbo.close()

    print("Finished")

elif par["test"] == "repartition":

    test_list = ["rad","rad+front","rad+back","front+back","front","back","rad+front+back"]

    A0 = []
    A1 = []
    A2 = []
    A3 = []
    A4 = []
    A6 = []
    A7 = []
    A8 = []

    for i in range(len(test_list)):

        if test_list[i] == "rad":
            par["eps"] = 0.9
            par["coeff_h_back"] = 1E-6
            par["a_htop"] = 1E-6
            par["b_htop"] = 1E-6
        elif test_list[i] == "rad+front":
            par["eps"] = 0.9
            par["coeff_h_back"] = 1E-6
            par["a_htop"] = 3.6
            par["b_htop"] = 3
        elif test_list[i] == "rad+back":
            par["eps"] = 0.9
            par["coeff_h_back"] = 1
            par["a_htop"] = 1E-6
            par["b_htop"] = 1E-6
        elif test_list[i] == "front+back":
            par["eps"] = 1E-6
            par["coeff_h_back"] = 1
            par["a_htop"] = 3.6
            par["b_htop"] = 3
        elif test_list[i] == "front":
            par["eps"] = 1E-6
            par["coeff_h_back"] = 1E-6
            par["a_htop"] = 3.6
            par["b_htop"] = 3
        elif test_list[i] == "back":
            par["eps"] = 1E-6
            par["coeff_h_back"] = 1
            par["a_htop"] = 1E-6
            par["b_htop"] = 1E-6
        elif test_list[i] == "rad+front+back":
            par["eps"] = 0.9
            par["coeff_h_back"] = 1
            par["a_htop"] = 3.6
            par["b_htop"] = 3

        # Model

        df_par,df,X = ty.test_meander_condi(par,G_list,coeff_G_p_list,T_amb_list,u_list,T_guess_list,T_f_in_list)

        A0.append(X[0][len(X[0])-1])
        A1.append(-X[0][0])
        A2.append(-X[0][1])
        A3.append(-X[0][2])
        A4.append(X[0][3])
        A6.append(-X[0][4])
        A7.append(-X[0][5])
        A8.append(-X[0][6])

    # Creating Excel outputs

    suffixe = r'\Parametric'+str(par["ailette"])+r'-'+moment+r'.xlsx'
    fichier_o = pathout+suffixe

    for d in dataframe_to_rows(df_par, index=True, header=False):
        sheet_o.append(d)

    wso = wbo.create_sheet("Outputs")

    wso['A'+str(1)] = 'test'
    wso['B'+str(1)] = 'a0'
    wso['C'+str(1)] = 'a1'
    wso['D'+str(1)] = 'a2'
    wso['E'+str(1)] = 'a3'
    wso['F'+str(1)] = 'a4'
    wso['G'+str(1)] = 'a6'
    wso['H'+str(1)] = 'a7'
    wso['I'+str(1)] = 'a8'
    wso['J'+str(1)] = 'a0 DS'
    wso['K'+str(1)] = 'a1 DS'


    for l in range(len(test_list)):
        wso['A'+str(l+2)] = test_list[l]
        wso['B'+str(l+2)] = A0[l]
        wso['C'+str(l+2)] = A1[l]
        wso['D'+str(l+2)] = A2[l]
        wso['E'+str(l+2)] = A3[l]
        wso['F'+str(l+2)] = A4[l]
        wso['G'+str(l+2)] = A6[l]
        wso['H'+str(l+2)] = A7[l]
        wso['I'+str(l+2)] = A8[l]
        wso['J'+str(l+2)] = A0[l]-A6[l]*(-1.7)
        wso['K'+str(l+2)] = A1[l]+A3[l]*(-1.7)

    wbo.save(filename = fichier_o)
    wbo.close()

    print("Finished")

elif par["test"] == "multi":

    
    # Creating Excel outputs
    suffixe = r'\Parametric'+str(par["ailette"])+r'-'+moment+r'.xlsx'
    fichier_o = pathout+suffixe

    wso = wbo.create_sheet("Outputs")


    eta_nom_list = [0.2]
    L_riser_list = [1.08]
    tau_alpha_list = [0.8]
    eps_list = [0.9]
    R_TOP_list = [0.004,0.005,0.006]
    a_htop_list = [3.2,3.5,3.85,4,4.2]
    coeff_htop_list = [1,1.2,1.4]
    coeff_h_back_list = [1,1.2,1.4]
    air_layer_list = np.linspace(0.00001,0.001,5)
    Cp_list = [3800,3900,4000,4100]
   
    A0 = []
    A1 = []
    A2 = []
    A3 = []
    A4 = []
    A6 = []
    A7 = []
    A8 = []

    total = len(eta_nom_list)*len(L_riser_list)*len(tau_alpha_list)*len(eps_list)*len(R_TOP_list)*len(a_htop_list)*len(coeff_htop_list)*len(coeff_h_back_list)*len(air_layer_list)*len(Cp_list)

    l=0

    for a in range(len(eta_nom_list)):
        for b in range(len(L_riser_list)):
            for c in range(len(tau_alpha_list)):
                for d in range(len(eps_list)):
                    for e in range(len(R_TOP_list)):
                        for f in range(len(a_htop_list)):
                            for g in range(len(coeff_htop_list)):
                                for h in range(len(coeff_h_back_list)):
                                    for i in range(len(air_layer_list)):
                                        for j in range(len(Cp_list)):

                                            print('test ',l, '/ ',total)

                                            par["eta_nom"] = eta_nom_list[a]
                                            par["L_riser"] = L_riser_list[b]
                                            par["tau_alpha"] = tau_alpha_list[c]
                                            par["eps"] = eps_list[d]
                                            par["R_TOP"] = R_TOP_list[e]
                                            par["a_htop_list"] = a_htop_list[f]
                                            par["coeff_b_htop"] = coeff_htop_list[g]
                                            par["coeff_h_back"] = coeff_h_back_list[h]
                                            par["air_layer"] = air_layer_list[i]
                                            ty.change_air_layer(par,air_layer_list[i])
                                            par["Cp"] = Cp_list[j]

                                            wso['A'+str(l+2)] = eta_nom_list[a]
                                            wso['B'+str(l+2)] = L_riser_list[b]
                                            wso['C'+str(l+2)] = tau_alpha_list[c]
                                            wso['D'+str(l+2)] = eps_list[d]
                                            wso['E'+str(l+2)] = R_TOP_list[e]
                                            wso['F'+str(l+2)] = a_htop_list[f]
                                            wso['G'+str(l+2)] =  coeff_htop_list[g]
                                            wso['H'+str(l+2)] = coeff_h_back_list[h]
                                            wso['I'+str(l+2)] = air_layer_list[i]
                                            wso['J'+str(l+2)] = Cp_list[j]

                                            # Model

                                            df_par,df,X = ty.test_meander_condi(par,G_list,coeff_G_p_list,T_amb_list,u_list,T_guess_list,T_f_in_list)

                                            A0.append(X[0][len(X[0])-1])
                                            A1.append(-X[0][0])
                                            A2.append(-X[0][1])
                                            A3.append(-X[0][2])
                                            A4.append(X[0][3])
                                            A6.append(-X[0][4])
                                            A7.append(-X[0][5])
                                            A8.append(-X[0][6])

                                            wso['K'+str(l+2)] = A0[l]
                                            wso['L'+str(l+2)] = A1[l]
                                            wso['M'+str(l+2)] = A2[l]
                                            wso['N'+str(l+2)] = A3[l]
                                            wso['O'+str(l+2)] = A4[l]
                                            wso['P'+str(l+2)] = A6[l]
                                            wso['Q'+str(l+2)] = A7[l]
                                            wso['R'+str(l+2)] = A8[l]
                                            wso['S'+str(l+2)] = A0[l]-A6[l]*(-1.7)
                                            wso['T'+str(l+2)] = A1[l]+A3[l]*(-1.7)

                                            l=l+1
    wso['A'+str(1)] = "eta_nom"
    wso['B'+str(1)] = "L_riser"
    wso['C'+str(1)] = "tau_alpha"
    wso['D'+str(1)] = "eps"
    wso['E'+str(1)] = "R_TOP"
    wso['F'+str(1)] = "a_htop"
    wso['G'+str(1)] = "b_htop"
    wso['H'+str(1)] = "coeff_h_back"
    wso['I'+str(1)] = "air_layer"
    wso['J'+str(1)] = "Cp"

    wso['K'+str(1)] = 'a0'
    wso['L'+str(1)] = 'a1'
    wso['M'+str(1)] = 'a2'
    wso['N'+str(1)] = 'a3'
    wso['O'+str(1)] = 'a4'
    wso['P'+str(1)] = 'a6'
    wso['Q'+str(1)] = 'a7'
    wso['R'+str(1)] = 'a8'
    wso['S'+str(1)] = 'a0 DS'
    wso['T'+str(1)] = 'a1 DS'


    sheet_o = wbo.create_sheet("Intputs")

    for d in dataframe_to_rows(df_par, index=True, header=False):
        sheet_o.append(d)

    wbo.save(filename = fichier_o)
    wbo.close()

    print("Finished")

# Test un panneau dans certaines conditions
# Copié dans l'ipynb
elif par["test"]=="champ":

    suffixe = r'\OnePanel-'+moment+r'.xlsx'
    fichier_o = pathout+suffixe
    var = {}

    list_var,list_var_conv = ty.meander(par,var,par["N_meander"],par["T_fluid_in0"],315,"all")

    compt = 2

    *var_names0, = var
    var_names = ['T_fluid_in'] + var_names0
    
    for m in range(len(var_names)):
        sheet_o.cell(row=1,column=m+1,value = var_names[m])

    for l in range(par["N_meander"]):
        for m in range(len(var_names)):
            sheet_o.cell(row=l+2,column=m+1,value = list_var[l][var_names[m]])

    sheet_o3 = wbo.create_sheet("Convergence")

    var_names2 = ['Slice','T_fluid_in'] + var_names0

    for m in range(len(var_names2)):
        sheet_o3.cell(row=1,column=m+1,value = var_names2[m])

    for l in range(len(list_var_conv)):
        for m in range(len(var_names2)):
            sheet_o3.cell(row=l+2,column=m+1,value = list_var_conv[l][var_names2[m]])
    

    sheet_o2 = wbo.create_sheet("T_abs")

    x_list = np.linspace(0,par["delta"],100)
    
    T_abs = []

    delta = par["delta"]

    for k in range(par["N_meander"]):
        b = list_var[k]["b"]
        j = list_var[k]["j"]
        m = list_var[k]["m"]
        T_B = list_var[k]["T_Base_mean"]

        b_j = b/j

        for l in range(len(x_list)):
            res = b_j+((T_B-b_j)/math.cosh(m*delta))*math.cosh(m*x_list[l])

            sheet_o2.cell(row=l+1,column=k+2,value = res)

    for l in range(len(x_list)):
        sheet_o2.cell(row=l+1,column=1,value=x_list[l])

    wbo.save(filename = fichier_o)
    wbo.close()

    print("Finished")
    
else:
    print("Choose a type of test")
